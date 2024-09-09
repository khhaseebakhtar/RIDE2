import random
import sys
import threading
import traceback
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles import Font


def make_file1(output, sheet_name, path, vendor, exe):
    physical_interfaces = output['physical interfaces']
    trunks = output['trunks']
    interface_descriptions = output['interface descriptions']
    inventory_pic_status = output['inventory pic status']
    licenses = output['licenses']
    optics = output['optics']

    version_info = output['version']
    inventory_details = output['inventory details']
    license_usage = output['license usage']
    port_license_usage = output['license usage on ports']

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    # Define headers and their positions
    headers = {
        'A1': 'INTERFACE', 'B1': 'PHY STATE', 'C1': 'ADMIN STATUS', 'D1': 'DESCRIPTION',
        'G1': 'PHYSICAL INTERFACE', 'H1': 'PHY STATE', 'I1': 'PORT BW', 'J1': 'LINK TYPE',
        'M1': 'TRUNK NAME', 'N1': 'TRUNK STATE', 'O1': 'TOTAL LINKS', 'P1': 'MEMBERS INTERFACES',
        'S1': 'LICENSE DESCRIPTION', 'T1': 'AVAILABLE LICENSES', 'U1': 'USED LICENSES',
        'V1': 'LICENSE EXPIRY', 'W1': 'ITEM NAME',
        'T20': 'SLOT', 'U20': 'SUB SLOT', 'V20': 'STATUS',
        'W20': 'TYPE', 'X20': 'PORT COUNT', 'Y20': 'PORT TYPE',
        'Z1': 'PORT', 'AA1': 'STATUS', 'AB1': 'DUPLEX',
        'AC1': 'TYPE', 'AD1': 'WL', 'AE1': 'RX POWER', 'AF1': 'TX POWER', 'AG1': 'MODE',
        'AI1': 'CHASSIS DETAILS', 'AI2': 'OS VERSION', 'AI3': 'SOFTWARE VERSION',
        'AI4': 'MODEL', 'AI5': 'UPTIME', 'AI6': 'SFU SLOTS', 'AI7': 'LPU SLOTS',
        'AI11': 'MODULE TYPE', 'AJ11': 'SLOT', 'AK11': 'BOARD TYPE', 'AL11': 'BAR CODE', 'AM11': 'DESCRIPTION'
    }

    wider_cells = [('Q', 'AA', 'AC', 'AJ', 'U', 'AJ', 10),
                   ('R', 'T',15),
                   ('A', 'G', 'T', 'M', 'I', 'J', 'AA', 'AC', 'AI', 'AL', 'AK', 20),
                   ('P','W', 'AM', 'AM', 30),
                   ('D', 'S', 50)]

    for each_change in wider_cells:
        for each_cell in range(0, each_change.__len__() - 1):
            worksheet.column_dimensions[each_change[each_cell]].width = each_change[-1]

    worksheet.column_dimensions['P'].hight = 10

    # Add vendor-specific headers & dictionary Keys
    if 'huawei' in vendor:
        header_change = [('Q1', 'MAX BW'), ('R1', 'AVAILABLE BW')]
        for change in header_change:
            headers[change[0]] = change[1]
        trunks_colmns = {'M': 'trunk_number', 'N': 'trunk_state', 'O': 'no_of_links', 'P': 'member_interfaces',
                         'Q': 'max_bw',
                         'R': 'current_bw'}
        optics_colmns = {'Z': 'port', 'AA': 'status', 'AB': 'duplex', 'AC': 'type', 'AD': 'wl', 'AE': 'rxpw',
                         'AF': 'txpw',
                         'AG': 'mode'}
        inventory_colmns = {'AI': 'module', 'AJ': 'slot_no', 'AK': 'board_type', 'AL': 'bar_code', 'AM': 'description'}
        green_fill_headers = ('G1', 'H1', 'I1', 'J1', 'Z1', 'AA1', 'AB1', 'AC1', 'AD1', 'AE1', 'AF1', 'AG1')

    if 'juniper' in vendor:
        header_change = [('AI6', 'ROUTING ENGINES'), ('AI7', 'FPC INSTALLED'),
                         ('AA1', 'TYPE'), ('AB1', 'WL'), ('AC1', 'MODE'), ('AJ11', 'PART NUMBER'),
                         ('AD1', ''), ('AE1', ''), ('AF1', ''), ('AG1', ''),
                         ('AJ11', 'VERSION'), ('AK11', 'PART NUMBER')]
        for change in header_change:
            headers[change[0]] = change[1]
        trunks_colmns = {'M': 'trunk_number', 'N': 'trunk_state', 'O': 'no_of_links', 'P': 'member_interfaces'}
        optics_colmns = {'Z': 'port', 'AA': 'type', 'AB': 'wl', 'AC': 'mode'}
        inventory_colmns = {'AI': 'module', 'AJ': 'version', 'AK': 'part_number', 'AL': 'bar_code', 'AM': 'description'}
        green_fill_headers = ('G1', 'H1', 'I1', 'J1', 'Z1', 'AA1', 'AB1', 'AC1')

    # Apply headers and their styles
    yellow_fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='00CCFFCC', end_color='00CCFFCC', fill_type='solid')
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    purpil_fill = PatternFill(start_color='00CC99FF', end_color='00CC99FF', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True)

    for cell, value in headers.items():
        worksheet[cell] = value
        worksheet[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet[cell].font = header_font
        if cell.startswith(('A1', 'B1', 'C1', 'D1', 'S1', 'T1', 'U1', 'V1', 'W1', 'S20', 'AI', 'AJ', 'AK', 'AL', 'AM')):
            worksheet[cell].fill = yellow_fill
        elif cell.startswith(green_fill_headers):
            worksheet[cell].fill = green_fill
        elif cell.startswith(('M1', 'N1', 'O1', 'P1', 'T20', 'U20', 'V20', 'W20', 'X20', 'Y20','Q1','R1')):
            worksheet[cell].fill = purpil_fill
        else:
            worksheet[cell].fill = red_fill

    # Write data to worksheet
    def write_data(data, start_row, columns, maker='huawei'):
        for row, item in enumerate(data, start=start_row):
            for col, key in columns.items():
                if col == 'N' and 'juniper' in maker:
                    worksheet[f"{col}{row}"] = f'=VLOOKUP(M{row},A:B,2,FALSE)'
                else:
                    worksheet[f"{col}{row}"] = item[key]

    write_data(interface_descriptions, 2, {'A': 'interface', 'B': 'phy', 'C': 'opr_status', 'D': 'description'})
    write_data(physical_interfaces, 2, {'G': 'interface', 'H': 'link_status', 'I': 'port_bw', 'J': 'type'})
    write_data(trunks, 2, trunks_colmns, vendor)
    # write_data(optics, 2,optics_colmns)
    # write_data(port_license_usage, 2, {'AC': 'port', 'AD': 'fname', 'AE': 'ncount', 'AF': 'ucount', 'AG': 'status'})
    # write_data(inventory_pic_status, 21,
    #      {'T': 'pic_slot', 'U': 'pic_sub', 'V': 'status', 'W': 'type', 'X': 'port_count', 'Y': 'port_type'})
    # write_data(inventory_details, 12,inventory_colmns)

    # for row, item in enumerate(licenses, start=2):
    #     cell = str(row)
    #     worksheet[f'S{cell}'] = item['description']
    #     worksheet[f'T{cell}'] = item['expired_date']
    #     for lic in license_usage:
    #         if lic['lic_name'].strip() == item['lic_name'].strip():
    #             worksheet[f'U{cell}'] = lic['avil_lic']
    #             worksheet[f'V{cell}'] = lic['used_lic']
    #             worksheet[f'W{cell}'] = lic['lic_name']
    #
    # worksheet['AJ2'] = version_info[0]['main_os']
    # worksheet['AJ3'] = version_info[0]['os_version']
    # worksheet['AJ4'] = version_info[0]['model']
    # worksheet['AJ5'] = version_info[0]['uptime']
    # worksheet['AJ6'] = version_info[0]['mpu_q']
    # worksheet['AJ7'] = version_info[0]['sru_q']
    # if 'huawei' in vendor:
    #     worksheet['AJ8'] = version_info[0]['sfu_q']
    #     worksheet['AJ9'] = version_info[0]['lpu_q']
    # if 'juniper' in vendor:
    #     worksheet['AJ6'] = '=COUNTIF(AI:AI, "Routing Engine *")'
    #     worksheet['AJ7'] = version_info[0]['lpu_q']

    # Save workbook
    try:
        workbook.save(f"{path}/{sheet_name}.xlsx")
    except PermissionError:
        file_name = f"{path}/{sheet_name}_{random.randint(0, 1000)}.xlsx"
        exe.sig.set_logging_signal.emit(f"{traceback.format_exc()}\n file is renamed to {file_name}")
        workbook.save(file_name)


class Writer():
    def __init__(self, output, device_name, path, vendor, exe):
        self.t = threading.Thread(args=(device_name, output, path, vendor, exe),
                                  target=make_file1(output, device_name, path, vendor, exe))
        self.t.start()
